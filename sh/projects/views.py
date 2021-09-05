
from django.shortcuts import render
from rest_framework import viewsets , permissions , serializers
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from .models import *
from django.contrib.auth.decorators import login_required
from taskBoard.models import task
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from django.http import HttpResponse
from rest_framework.response import Response
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill , Font
import string
from django.db.models import Q,Sum,F
import json
import os
import pandas as pd
import re
from django.conf import settings as globalSettings
import datetime
from dateutil.relativedelta import relativedelta
from HR.models import Team,designation
from django.core.exceptions import PermissionDenied,SuspiciousOperation
from django.shortcuts import render, redirect , get_object_or_404
import openpyxl
import sys, traceback
reload(sys)
sys.setdefaultencoding("utf-8")

from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.http import JsonResponse
from tools.models import *
import zipfile
from django.core.exceptions import ObjectDoesNotExist , SuspiciousOperation
from django.contrib.auth import authenticate , login

def cleanStr(val):
    try:
        return str(val.encode("utf-8"))
    except:
        return str(val)

# Create your views here.
class PlannerItemViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = PlannerItemSerializer
    queryset = PlannerItem.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['company' , 'teamLead' , 'location']

class mediaViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = mediaSerializer
    queryset = media.objects.all()

class projectCommentViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = projectCommentSerializer
    queryset = projectComment.objects.all()

class projectViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = projectSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title','projectClosed','company','status']
    def get_queryset(self):
        if 'LDDprojects' in self.request.GET and 'userId' in self.request.GET:
            userObj = User.objects.get(pk=int(self.request.GET['userId']))
            projectsPkList = list(set(list(LDDAssignment.objects.filter(Q(coder=userObj)|Q(qc=userObj)).values_list('project',flat=True).distinct())))
            return project.objects.filter(~Q(status='complete'),pk__in=projectsPkList).order_by('-created')
        if 'userProjectsOnly' in self.request.GET and 'userPk' in self.request.GET:
            userObj = User.objects.get(pk=int(self.request.GET['userPk']))
            projectsPkList = list(set(list(FileUpload.objects.filter(Q(coder=userObj)|Q(qa=userObj)|Q(recoder=userObj)|Q(reqa=userObj)).values_list('project',flat=True).distinct())))
            return project.objects.filter(pk__in=projectsPkList).order_by('-created')
        u = self.request.user
        if u.is_superuser:
            if 'dashboardProjects' in self.request.GET:
                return project.objects.filter(~Q(status='complete'),disabled=False).order_by('-created')
            else:
                return project.objects.all().order_by('-created')
        else:
            userTams = list(u.designation.team.all().values_list('pk',flat=True))
            toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) | Q(coderTeam__in = userTams) | Q(qaTeam__in = userTams) | Q(reCoderTeam__in = userTams) | Q(reQaTeam__in = userTams))
            if 'dashboardProjects' in self.request.GET:
                return toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')
            else:
                return toRet.distinct().order_by('-created')

class projectLiteListViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = projectLiteListSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title','projectClosed','company','status']
    def get_queryset(self):
        if 'LDDprojects' in self.request.GET and 'userId' in self.request.GET:
            userObj = User.objects.get(pk=int(self.request.GET['userId']))
            projectsPkList = list(set(list(LDDAssignment.objects.filter(Q(coder=userObj)|Q(qc=userObj)).values_list('project',flat=True).distinct())))
            return project.objects.filter(~Q(status='complete'),pk__in=projectsPkList).order_by('-created')
        if 'userProjectsOnly' in self.request.GET and 'userPk' in self.request.GET:
            userObj = User.objects.get(pk=int(self.request.GET['userPk']))
            projectsPkList = list(set(list(FileUpload.objects.filter(Q(coder=userObj)|Q(qa=userObj)|Q(recoder=userObj)|Q(reqa=userObj)).values_list('project',flat=True).distinct())))
            return project.objects.filter(pk__in=projectsPkList).order_by('-created')
        u = self.request.user
        if u.is_superuser:
            if 'dashboardProjects' in self.request.GET:
                return project.objects.filter(~Q(status='complete'),disabled=False).order_by('-created')
            else:
                return project.objects.all().order_by('-created')
        else:
            userTams = list(u.designation.team.all().values_list('pk',flat=True))
            toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) | Q(coderTeam__in = userTams) | Q(qaTeam__in = userTams) | Q(reCoderTeam__in = userTams) | Q(reQaTeam__in = userTams))
            if 'dashboardProjects' in self.request.GET:
                return toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')
            else:
                return toRet.distinct().order_by('-created')


class projectDashboardViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = projectDashboardLiteSerializer
    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['title','projectClosed','company','status']
    def get_queryset(self):
        u = self.request.user
        if u.is_superuser:
            if 'dashboardProjects' in self.request.GET:
                return project.objects.filter(~Q(status='complete'),disabled=False).order_by('-created')
            else:
                return project.objects.all().order_by('-created')
        else:
            userTams = list(u.designation.team.all().values_list('pk',flat=True))
            toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) | Q(coderTeam__in = userTams) | Q(qaTeam__in = userTams) | Q(reCoderTeam__in = userTams) | Q(reQaTeam__in = userTams))
            if 'dashboardProjects' in self.request.GET:
                return toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')
            else:
                return toRet.distinct().order_by('-created')

class projectLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, permissions.DjangoModelPermissionsOrAnonReadOnly)
    serializer_class = projectLiteSerializer
    queryset = project.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title','company']

class timelineItemViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = timelineItemSerializer
    queryset = timelineItem.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['project', 'category']

class ProjectFieldsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ProjectFieldsSerializer
    queryset = ProjectFields.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['project', 'dynamicForm']


class OCRConfigViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = OCRConfigSerializer
    queryset = OCRConfig.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['owner', 'name']

import random

class FileUploadViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = FileUploadSerializer
    # queryset = FileUpload.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['project','parent', 'id']
    def get_queryset(self):
        if 'getNextFile' in self.request.GET and 'projectPk' in self.request.GET and 'statusType' in self.request.GET:
            projObj = project.objects.get(pk=int(self.request.GET['projectPk']))
            toRet = FileUpload.objects.filter(~Q(status='approved'),project_id=int(self.request.GET['projectPk']),parent__isnull=True)
            # return toRet
            if self.request.GET['statusType'] == 'coder':
                notCompleteFiles = toRet.filter(coder=self.request.user,coded=False)
                if notCompleteFiles.exists():
                    return notCompleteFiles[0:1]
                else:
                    newFiles = toRet.filter(locked=False)
                    if newFiles.exists():

                        toPick = random.randint(0,newFiles.count())
                        # print "toPick" , toPick
                        # print " newFiles.count() : " , newFiles.count()
                        if toPick == newFiles.count():
                            toPick += -1

                        toRet = newFiles[toPick:toPick+1]
                        obj = toRet[0]
                        # print "file upload obj : " , obj
                        obj.locked=True
                        obj.relocked = True
                        obj.coder=self.request.user
                        obj.recoder=self.request.user
                        obj.save()
                        return FileUpload.objects.filter(pk = obj.pk )
                    else:
                        return newFiles
            elif self.request.GET['statusType'] == 'qa':
                notCompleteReviewFiles = toRet.filter(qa=self.request.user,checked=False,coded=True)
                if notCompleteReviewFiles.exists():
                    return notCompleteReviewFiles[0:1]
                else:
                    newReviewFiles = toRet.filter(checklocked=False,coded=True)
                    if newReviewFiles.exists():
                        # data:{checklocked:true,updateqaVal:true,rechecklocked:true,updatereqaVal:true}
                        toPick = random.randint(0,newReviewFiles.count())
                        if toPick == newReviewFiles.count():
                            toPick += -1
                        toRet = newReviewFiles[toPick:toPick+1]
                        print "toPick" , toPick
                        print " newFiles.count() : " , newReviewFiles.count()

                        obj = toRet[0]
                        # print "file upload obj : " , obj
                        obj.checklocked=True
                        obj.rechecklocked = True
                        obj.reqa=self.request.user
                        obj.qa=self.request.user
                        obj.save()
                        return FileUpload.objects.filter(pk = obj.pk )
                    else:
                        return newReviewFiles
            elif self.request.GET['statusType'] == 'recoding':
                notCompleteReviewFiles = toRet.filter(recoder=self.request.user,recoded=False,checked=True)
                if notCompleteReviewFiles.exists():
                    toRet = notCompleteReviewFiles[0:1]
                else:
                    toRet = notCompleteReviewFiles

                obj = toRet[0]
                obj.relocked=True
                obj.recoder=self.request.user
                obj.save()
                return toRet

            elif self.request.GET['statusType'] == 'reqa':
                notCompleteReviewFiles = toRet.filter(reqa=self.request.user,rechecked=False,recoded=True)
                if notCompleteReviewFiles.exists():
                    return notCompleteReviewFiles[0:1]
                else:
                    return notCompleteReviewFiles
                obj = toRet[0]
                obj.rechecklocked=True
                obj.reqa=self.request.user
                obj.save()

                return toRet

            else:
                return toRet
        if 'parentFiles' in self.request.GET:
            return FileUpload.objects.filter(parent__isnull=True)
        if 'pendingFiles' in self.request.GET:
            toRet = FileUpload.objects.filter(~Q(status='approved'),parent__isnull=True)
            if 'user' in self.request.GET:
                if self.request.GET.get('assignTyp','') == 'recoding':
                    toRet = toRet.filter(coded=True,recoded=False)
                    if int(self.request.GET['user'])>0:
                        toRet = toRet.filter(coder=int(self.request.GET['user']))
                elif self.request.GET.get('assignTyp','') == 'reqc':
                    toRet = toRet.filter(checked=True,rechecked=False)
                    if int(self.request.GET['user'])>0:
                        toRet = toRet.filter(qa=int(self.request.GET['user']))
            else:
                if self.request.GET.get('assignTyp','') == 'coding':
                    toRet = toRet.filter(coded=False)
                elif self.request.GET.get('assignTyp','') == 'qc':
                    toRet = toRet.filter(checked=False,coded=True)

            return toRet
        return FileUpload.objects.all()

class GetCoderAndQcDataAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        toRet = {'codersList':[],'qcList':[],'recodersList':[],'reqcList':[],'canCodeList':[],'canQcList':[]}
        # print 'req dataaaaaaaa',self.request.GET
        projectObj = project.objects.get(pk=int(self.request.GET['project']))
        filesObj = FileUpload.objects.filter(project=projectObj)
        codersData = list(set(list(filesObj.filter(coder__isnull=False).values_list('coder',flat=True))))
        qcData = list(set(list(filesObj.filter(qa__isnull=False).values_list('qa',flat=True))))
        # print codersData,qcData
        for i in codersData:
            userObj = User.objects.get(pk=int(i))
            name = userObj.first_name
            if userObj.last_name:
                name += ' ' + userObj.last_name
            toRet['codersList'].append({'pk':userObj.pk,'name':name})
        for i in qcData:
            userObj = User.objects.get(pk=int(i))
            name = userObj.first_name
            if userObj.last_name:
                name += ' ' + userObj.last_name
            toRet['qcList'].append({'pk':userObj.pk,'name':name})
        recodersList = list(set(list(projectObj.reCoderTeam.teamMembers.all().values_list('user',flat=True))))
        for i in recodersList:
            userObj = User.objects.get(pk=int(i))
            name = userObj.first_name
            if userObj.last_name:
                name += ' ' + userObj.last_name
            toRet['recodersList'].append({'pk':userObj.pk,'name':name})
        reqcList = list(set(list(projectObj.reQaTeam.teamMembers.all().values_list('user',flat=True))))
        for i in reqcList:
            userObj = User.objects.get(pk=int(i))
            name = userObj.first_name
            if userObj.last_name:
                name += ' ' + userObj.last_name
            toRet['reqcList'].append({'pk':userObj.pk,'name':name})
        canCodeList = list(set(list(projectObj.coderTeam.teamMembers.all().values_list('user',flat=True))))
        for i in canCodeList:
            userObj = User.objects.get(pk=int(i))
            name = userObj.first_name
            if userObj.last_name:
                name += ' ' + userObj.last_name
            toRet['canCodeList'].append({'pk':userObj.pk,'name':name})
        canQcList = list(set(list(projectObj.qaTeam.teamMembers.all().values_list('user',flat=True))))
        for i in canQcList:
            userObj = User.objects.get(pk=int(i))
            name = userObj.first_name
            if userObj.last_name:
                name += ' ' + userObj.last_name
            toRet['canQcList'].append({'pk':userObj.pk,'name':name})
        return Response(toRet,status=status.HTTP_200_OK)

class AssignFilesToUserAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        # print 'postttttttt req dataaaaaaaa',self.request.data
        filesObj = FileUpload.objects.filter(pk__in=self.request.data['filesPks'])
        userObj = User.objects.get(pk=int(self.request.data['userPk']))
        if self.request.data['typ'] == 'coding':
            filesObj.update(locked=True,coder=userObj)
        if self.request.data['typ'] == 'qc':
            filesObj.update(checklocked=True,qa=userObj)
        if self.request.data['typ'] == 'recoding':
            filesObj.update(relocked=True,recoder=userObj)
        if self.request.data['typ'] == 'reqc':
            filesObj.update(rechecklocked=True,reqa=userObj)
        return Response({},status=status.HTTP_200_OK)


class UnlockFilesForUserAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        # print 'postttttttt req dataaaaaaaa',self.request.data
        filesObj = FileUpload.objects.filter(pk__in=self.request.data['filesPks'])
        if self.request.data['typ'] == 'coding':
            filesObj.update(locked=False,coder=None, relocked = False , recoder = None)
        if self.request.data['typ'] == 'qc':
            filesObj.update(checklocked=False,qa=None , rechecklocked = False , reqa = None)
        if self.request.data['typ'] == 'recoding':
            filesObj.update(relocked=False,recoder=None)
        if self.request.data['typ'] == 'reqc':
            filesObj.update(rechecklocked=False,reqa=None)

        return Response({},status=status.HTTP_200_OK)




class DashboardDataProjectMgmtAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        toRet = {}
        # print ' Dashboard Data ProjectMgmt'
        projectObj = project.objects.filter(projectClosed=False)
        teamCount = 0
        for i in projectObj:
            pTCount = i.team.all().count()
            teamCount += pTCount
        issue = Issues.objects.all()
        # taskObj = task.objects.all()
        toRet['projectCount']=projectObj.count()
        toRet['issueCount']=issue.count()
        toRet['teamCount']=teamCount
        # toRet['taskCount']=taskObj.count()
        return Response(toRet,status=status.HTTP_200_OK)

class CheckNextFileAPI(APIView):
    def get(self, request, format=None):
        toReturn = {'nxt':False}
        # print request.GET,'getttttttttttt'
        if 'typ' in request.GET and 'projectPk' in request.GET:
            if request.GET['typ'] == 'coder':
                toRet = FileUpload.objects.filter(~Q(status='approved'),project_id=int(self.request.GET['projectPk']),parent__isnull=True,coder=self.request.user,coded=False)
                if toRet.exists():
                    # print 'coding existssss'
                    toReturn['nxt'] = True
            elif request.GET['typ'] == 'recoder':
                unlcokedDocs = FileUpload.objects.filter(~Q(status='approved'),project_id=int(self.request.GET['projectPk']),parent__isnull=True,recoder__isnull=True,recoded=False,checked=True)

                if unlcokedDocs.count()>0:
                    # print "available unlocked documents"
                    obj = unlcokedDocs[0]
                    # print "file upload obj : " , obj
                    obj.relocked = True
                    obj.recoder=self.request.user
                    obj.save()
                    return Response({"nxt": True} , status=status.HTTP_200_OK)

                toRet = FileUpload.objects.filter(~Q(status='approved'),project_id=int(self.request.GET['projectPk']),parent__isnull=True,recoder=self.request.user,recoded=False,checked=True)
                # here check if any other unlocked recoding is available , lock it and return



                if toRet.exists():
                    # print 'recoding existssss'
                    toReturn['nxt'] = True
            elif request.GET['typ'] == 'qc':
                toRet = FileUpload.objects.filter(~Q(status='approved'),project_id=int(self.request.GET['projectPk']),parent__isnull=True,qa=self.request.user,checked=False,coded=True)
                if toRet.exists():
                    # print 'qc existssss'
                    toReturn['nxt'] = True
            elif request.GET['typ'] == 'reqc':

                unlcokedDocs = FileUpload.objects.filter(~Q(status='approved'),project_id=int(self.request.GET['projectPk']),parent__isnull=True,reqa__isnull=True,rechecked=False,recoded=True)

                if unlcokedDocs.count()>0:
                    # print "available unlocked documents"
                    obj = unlcokedDocs[0]
                    # print "file upload obj : " , obj
                    obj.rechecklocked = True
                    obj.reqa=self.request.user
                    obj.save()
                    return Response({"nxt": True} , status=status.HTTP_200_OK)


                toRet = FileUpload.objects.filter(~Q(status='approved'),project_id=int(self.request.GET['projectPk']),parent__isnull=True,reqa=self.request.user,rechecked=False,recoded=True)
                if toRet.exists():
                    # print 'reqc existssss'
                    toReturn['nxt'] = True

        return Response(toReturn,status=status.HTTP_200_OK)

class FieldValueViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = FieldValueSerializer
    queryset = FieldValue.objects.all().order_by('order')
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['field','file' , 'project','needCorrection']

class LDDAssignmentViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = LDDAssignmentSerializer
    # queryset = LDDAssignment.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['project','coder' , 'qc', 'coded', 'checked', 'updated']
    def get_queryset(self):
        if 'coderorqc' in self.request.GET:
            userObj = User.objects.get(pk=int(self.request.GET['coderorqc']))
            return LDDAssignment.objects.filter(Q(coder=userObj) | Q(qc=userObj))
        else:
            return LDDAssignment.objects.all()


class FileUploadTracerViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = FileUploadTracerSerializer
    # queryset = FileUpload.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['project' , 'path']
    def get_queryset(self):
        if 'user' in self.request.GET:
            userObj = User.objects.get(pk=int(self.request.GET['user']))
            return FileUpload.objects.filter(Q(coder=userObj) | Q(qa=userObj) , parent = None)
        else:
            return FileUpload.objects.all()

class ProjectPerformaceReportAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        workbook = Workbook()
        Sheet1 = workbook.active
        projectObj = project.objects.get(pk = request.GET['projectpk'])
        Sheet1.title = projectObj.title
        fieldPkList = list(ProjectFields.objects.filter(project = projectObj).values_list('pk',flat=True))
        hdFont = Font(size=12,bold=True)
        red_font = Font(color='00e62929')
        alphaChars = list(string.ascii_uppercase)

        projectFieldObj = ProjectFields.objects.filter(project = projectObj,pk__in=fieldPkList)
        hd = ['Begin Doc','End Doc','Coder','QC','Re Coder','Re Qc', 'Coding Time' , 'QC Time' , 'Re Coding Time' , 'Re QC Time' , 'Total Time']
        flds = list(projectFieldObj.values_list('name',flat=True))
        hd += flds
        Sheet1.append(hd)
        hdWidth = []
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            hdWidth.append(15)

        fieldValues = FieldValue.objects.filter(project = projectObj,field__in=fieldPkList)

        totalProjectFiles = FileUpload.objects.filter(project = projectObj,parent__isnull=True)

        for flIdx,fl in enumerate(totalProjectFiles):
            hd1 = []
            wCl = []
            if fl.path:
                fPathVal = str(fl.path).split('/')[-1]
            else:
                fPathVal = str(fl.file).split('/')[-1]
                if '__' in fPathVal:
                    fPathVal = fPathVal.split('__')[-1]
            hd1.append(fPathVal)
            if hdWidth[0]<=len(fPathVal):
                hdWidth[0] = len(fPathVal)+5


            chld = FileUpload.objects.filter(project = projectObj,parent=fl).order_by('pk')
            # print "================"
            # print fPathVal
            # for cld in chld:
            #     print cld.path

            if chld.exists():
                chd = chld.last()
                if chd.path:
                    fPathVal2 = str(chd.path).split('/')[-1]
                else:
                    fPathVal2 = str(chd.file).split('/')[-1]
                    if '__' in fPathVal2:
                        fPathVal2 = fPathVal2.split('__')[-1]
            else:
                fPathVal2 = hd1[0]
            hd1.append(fPathVal2)
            if hdWidth[1]<=len(fPathVal2):
                hdWidth[1] = len(fPathVal2)+10

            if fl.coder:
                hd1.append(fl.coder.first_name)
                if hdWidth[2]<=len(fl.coder.first_name):
                    hdWidth[2] = len(fl.coder.first_name)+10
            else:
                hd1.append('')
            if fl.qa:
                hd1.append(fl.qa.first_name)
                if hdWidth[3]<=len(fl.qa.first_name):
                    hdWidth[3] = len(fl.qa.first_name)+10
            else:
                hd1.append('')
            if fl.recoder:
                hd1.append(fl.recoder.first_name)
                if hdWidth[4]<=len(fl.recoder.first_name):
                    hdWidth[4] = len(fl.recoder.first_name)+10
            else:
                hd1.append('')
            if fl.reqa:
                hd1.append(fl.reqa.first_name)
                if hdWidth[5]<=len(fl.reqa.first_name):
                    hdWidth[5] = len(fl.reqa.first_name)+5
            else:
                hd1.append('')

            hd1.append( str(fl.codeTime) )
            hd1.append( str(fl.qcTime) )
            hd1.append(str(fl.reCodeTime))
            hd1.append(str(fl.reQcTime))
            hd1.append(str(fl.reQcTime + fl.codeTime + fl.qcTime + fl.reCodeTime ))



            fileFiledValues = fieldValues.filter(file=fl)
            for ridx,i in enumerate(fieldPkList):
                val = ''
                objs = fieldValues.filter(file=fl,field_id=i)
                needCorr = [n.reCodedField for n in objs]
                for sepIdx,j in enumerate(objs):
                    # print objs , "field .........."
                    if j.char and (j.field.type == 'dropdown' or j.field.type == 'smallText'):
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += cleanStr(j.char)
                    if j.dt and j.field.type == 'date':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        if j.field.dateFormat and len(j.field.dateFormat)>0:
                            try:
                                dft = datetime.datetime.strftime(j.dt, str(j.field.dateFormat))
                            except:
                                pass
                        else:
                            dft = j.dt
                        val += str(dft)
                    if j.dtTime and j.field.type == 'datetime':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        if j.field.dateFormat and len(j.field.dateFormat)>0:
                            try:
                                dft = datetime.datetime.strftime(j.dtTime, str(j.field.dateFormat)+' %H:%M:%S')
                            except:
                                pass
                        else:
                            dft = j.dtTime
                        val += str(dft)
                    if j.txt and j.field.type == 'text':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += cleanStr(j.txt)
                    if j.field.type == 'boolean':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += str(j.bool)
                    if j.email and j.field.type == 'email':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += cleanStr(j.email)
                    if j.field.type == 'name':

                        name = ''
                        if j.prefix:
                            prefix = cleanStr(j.prefix)
                        else:
                            prefix = ''
                        if j.fname:
                            fname = cleanStr(j.fname)
                        else:
                            fname = ''
                        if j.mName:
                            mname = cleanStr(j.mName)
                        else:
                            mname = ''
                        if j.lname:
                            lname = cleanStr(j.lname)
                        else:
                            lname = ''
                        if j.compName:
                            compname = cleanStr(j.compName)
                        else:
                            compname = ''
                        nstrtemp = "".join([fname , mname , lname , compname , prefix])
                        if len(nstrtemp) == 0:
                            continue
                        j.field.nameFormat = j.field.nameFormat.replace('\\s' , ' ')
                        name = j.field.nameFormat.replace('FN' , fname).replace('MN' , mname).replace('LN' , lname).replace('PN' , prefix).replace('|' , '').replace('CN' , compname).strip()
                        if name == "":
                            continue

                        nfData = str(j.field.nameFormat).split('|')
                        nfData.append(',')
                        for repeat in [0,1,2,3]:
                            name = cleanName(name , nfData).strip()
                        name = ' '.join(name.split())

                        if nfData[7] in name:
                            cIndx = name.index(nfData[7])
                            if len(name)>2 and name[cIndx-2] != " " and name[cIndx-2] in nfData:
                                inxRoRemove = cIndx -2
                                if len(name) > inxRoRemove:
                                    name = name[0 : inxRoRemove : ] + name[inxRoRemove + 1 : :]
                            cIndx = name.index(nfData[7])
                            if name[cIndx-1] == " ":
                                inxRoRemove = cIndx -1
                                if len(name) > inxRoRemove:
                                    name = name[0 : inxRoRemove : ] + name[inxRoRemove + 1 : :]

                        if j.field.multiSeperator is None:
                            j.field.multiSeperator = ";"

                        if val == ",":
                            continue

                        if val == "":
                            val = name
                        else:
                            val += j.field.multiSeperator +  name

                if hdWidth[ridx+11]<=len(val):
                    hdWidth[ridx+11] = len(val)+5

                if len(val)>0 and any(needCorr):
                    wCl.append(str(alphaChars[ridx+6])+str(flIdx+2))

                hd1.append(val)

            Sheet1.append(hd1)
            # print wCl,'wClwCl'
            for wr in wCl:
                Sheet1[wr].font = red_font
            # print hdWidth,'col widthhhhh'
        for idx,i in enumerate(hdWidth):
            Sheet1.column_dimensions[alphaChars[idx]].width = i

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Project.xlsx'
        return response

def cleanName(name , nfData):
    for nidx , nf in enumerate(nfData):
        if nf == '' or nf == " ":
            continue
        if name.startswith(nf):
            name = name[1:].strip()
            # print "starts with : " , nf , len(nf) , " so replacing it : ", name
        if name.endswith(nf):
            name = name[0:-1].strip()
            # print "ends with : " , nf , " so replacing it : ", name
    return name


class ProjectCCViewAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        projectObj = project.objects.get(pk = request.GET['projectpk'])
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        projectFieldObj = ProjectFields.objects.filter(project = projectObj)
        hd = [ 'Begin Doc','End Doc']
        flds = list(projectFieldObj.values_list('name',flat=True))
        hd += flds
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
        hd.append('ID')
        toRet = [hd]
        fieldValues = FieldValue.objects.filter(project = projectObj)

        totalProjectFiles = FileUpload.objects.filter(project = projectObj,parent__isnull=True ).all()[int(request.GET['offset']) :int(request.GET['offset']) + int(request.GET['limit'])]

        for fl in totalProjectFiles:
            hd1 = []
            if fl.path:
                fPathVal = str(fl.path).split('/')[-1]
            else:
                fPathVal = str(fl.file).split('/')[-1]
                if '__' in fPathVal:
                    fPathVal = fPathVal.split('__')[-1]
            hd1.append(fPathVal)


            chld = FileUpload.objects.filter(project = projectObj,parent=fl).order_by('pk')
            # print "================"
            # print fPathVal
            # for cld in chld:
            #     print cld.path
            if chld.exists():
                chd = chld.last()
                if chd.path:
                    fPathVal2 = str(chd.path).split('/')[-1]
                else:
                    fPathVal2 = str(chd.file).split('/')[-1]
                    if '__' in fPathVal2:
                        fPathVal2 = fPathVal2.split('__')[-1]
            else:
                fPathVal2 = hd1[0]
            hd1.append(fPathVal2)

            fileFiledValues = fieldValues.filter(file=fl)
            for ridx,i in enumerate(projectFieldObj):
                val = ''
                objs = fieldValues.filter(file=fl,field=i)
                for sepIdx,j in enumerate(objs):
                    # print j , "field .........." , j.field.type , j

                    if j.char and (j.field.type == 'dropdown' or j.field.type == 'smallText'):
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += cleanStr(j.char)
                    if j.dt and j.field.type == 'date':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        if j.field.dateFormat and len(j.field.dateFormat)>0:
                            try:
                                dft = datetime.datetime.strftime(j.dt, str(j.field.dateFormat))
                            except:
                                pass
                        else:
                            dft = j.dt
                        val += str(dft)
                    if j.dtTime and j.field.type == 'datetime':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        if j.field.dateFormat and len(j.field.dateFormat)>0:
                            try:
                                dft = datetime.datetime.strftime(j.dtTime, str(j.field.dateFormat)+' %H:%M:%S')
                            except:
                                pass
                        else:
                            dft = j.dtTime
                        val += str(dft)
                    if j.txt and j.field.type == 'text':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += cleanStr(j.txt)
                    if j.field.type == 'boolean':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += str(j.bool)
                    if j.email and j.field.type == 'email':
                        if sepIdx>0 and j.field.multiSeperator:
                            val += str(j.field.multiSeperator)
                        val += cleanStr(j.email)
                    if j.field.type == 'name':
                        name = ''
                        if j.prefix:
                            prefix = cleanStr(j.prefix).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                        else:
                            prefix = ''
                        if j.fname:
                            fname = cleanStr(j.fname).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                        else:
                            fname = ''
                        if j.mName:
                            mname = cleanStr(j.mName).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                        else:
                            mname = ''
                        if j.lname:
                            lname = cleanStr(j.lname).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                        else:
                            lname = ''
                        if j.compName:
                            compname = cleanStr(j.compName).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                        else:
                            compname = ''
                        nstrtemp = "".join([fname , mname , lname , compname , prefix])
                        if len(nstrtemp) == 0:
                            continue
                        j.field.nameFormat = j.field.nameFormat.replace('\\s' , ' ')
                        name = j.field.nameFormat.replace('FN' , fname).replace('MN' , mname).replace('LN' , lname).replace('PN' , prefix).replace('|' , '').replace('CN' , compname).strip()
                        if name == "":
                            continue

                        nfData = str(j.field.nameFormat).split('|')
                        nfData.append(',')
                        for repeat in [0,1,2,3]:
                            name = cleanName(name , nfData).strip()
                        name = ' '.join(name.split())

                        if nfData[7] in name:
                            cIndx = name.index(nfData[7])
                            if name[cIndx-2] != " " and name[cIndx-2] in nfData:
                                inxRoRemove = cIndx -2
                                if len(name) > inxRoRemove:
                                    name = name[0 : inxRoRemove : ] + name[inxRoRemove + 1 : :]
                            cIndx = name.index(nfData[7])
                            if name[cIndx-1] == " ":
                                inxRoRemove = cIndx -1
                                if len(name) > inxRoRemove:
                                    name = name[0 : inxRoRemove : ] + name[inxRoRemove + 1 : :]
                        name = name.replace( '__', 'PN').replace( '**', 'FN').replace( '^^', 'MN' ).replace( '@@', 'LN').replace( '##', 'CN' )
                        if j.field.multiSeperator is None:
                            j.field.multiSeperator = ";"

                        if val == ",":
                            continue

                        if val == "":
                            val = name
                        else:
                            val += j.field.multiSeperator +  name

                hd1.append(val)
            hd1.append(fl.pk)
            toRet.append(hd1)
        return Response(toRet ,status=status.HTTP_200_OK)



def projectReportExcel(request, forceRaw):
    workbook = Workbook()
    Sheet1 = workbook.active
    projectObj = project.objects.get(pk = request.GET['projectpk'])
    Sheet1.title = projectObj.title

    if 'pkList' in request.GET:
        selectedFields = json.loads(request.GET['pkList'])
        fieldPkList = [i['pk'] for i in selectedFields if i['selected']]

        print "fieldPkList : " , fieldPkList

        projectFieldObj = ProjectFields.objects.filter(project = projectObj,pk__in=fieldPkList)
    else:

        projectFieldObj = ProjectFields.objects.filter(project = projectObj)
        fieldPkList = [pfo.pk for pfo in projectFieldObj]



    hdFont = Font(size=12,bold=True)
    alphaChars = list(string.ascii_uppercase)

    hd = ['Begin Doc','End Doc']
    flds = list(projectFieldObj.values_list('name',flat=True))
    hd += flds
    Sheet1.append(hd)
    hdWidth = []
    for idx,i in enumerate(hd):
        cl = str(alphaChars[idx])+'1'
        Sheet1[cl].font = hdFont
        hdWidth.append(10)

    if 'pkList' in request.GET:
        fieldValues = FieldValue.objects.filter(project = projectObj,field__in=fieldPkList)
    else:
        fieldValues = FieldValue.objects.filter(project = projectObj)

    if 'raw' in request.GET or forceRaw:
        totalProjectFiles = FileUpload.objects.filter(project = projectObj,parent__isnull=True , coded = True)
        # print "raw : ", totalProjectFiles.count()
    else:
        totalProjectFiles = FileUpload.objects.filter(project = projectObj,parent__isnull=True , status = 'approved')
        # print "non raw : ", totalProjectFiles.count()

    for fl in totalProjectFiles:
        hd1 = []
        if fl.path:
            fPathVal = str(fl.path).split('/')[-1]
        else:
            fPathVal = str(fl.file).split('/')[-1]
            if '__' in fPathVal:
                fPathVal = fPathVal.split('__')[-1]
        hd1.append(fPathVal)
        if hdWidth[0]<=len(fPathVal):
            hdWidth[0] = len(fPathVal)+15

        chld = FileUpload.objects.filter(project = projectObj,parent=fl).order_by('pk')

        if chld.exists():
            chd = chld.last()
            if chd.path:
                fPathVal2 = str(chd.path).split('/')[-1]
            else:
                fPathVal2 = str(chd.file).split('/')[-1]
                if '__' in fPathVal2:
                    fPathVal2 = fPathVal2.split('__')[-1]
        else:
            fPathVal2 = hd1[0]
        hd1.append(fPathVal2)
        if hdWidth[1]<=len(fPathVal2):
            hdWidth[1] = len(fPathVal2)+15

        fileFiledValues = fieldValues.filter(file=fl)
        for ridx,i in enumerate(fieldPkList):
            val = ''
            objs = fieldValues.filter(file=fl,field_id=i)
            for sepIdx,j in enumerate(objs):


                if j.char and (j.field.type == 'dropdown' or j.field.type == 'smallText'):
                    if sepIdx>0 and j.field.multiSeperator:
                        val += str(j.field.multiSeperator)
                    try:
                        val += str(j.char)
                    except:
                        val += str(j.char.encode("utf-8"))

                if j.dt and j.field.type == 'date':
                    dft = ""
                    if sepIdx>0 and j.field.multiSeperator:
                        val += str(j.field.multiSeperator)

                    if j.field.dateFormat and len(j.field.dateFormat)>0:
                        try:
                            dft = datetime.datetime.strftime(j.dt, str(j.field.dateFormat))
                        except:
                            dft = "%s-%s-%s"%(j.dt.month ,j.dt.day, j.dt.year)
                    else:
                        dft = j.dt
                    if 'Month&Year' == j.dtMode:
                        try:
                            dft = datetime.datetime.strftime(j.dt,  "%m-00-%Y" )
                        except:
                            dft = "%s-00-%s"%(j.dt.month , j.dt.year)

                    elif 'Only Year' == j.dtMode:
                        dft = '00-00-' + str(j.dt.year)

                    elif 'date&month' == j.dtMode:
                        if j.dt is not None:
                            dft = datetime.datetime.strftime(j.dt, '%m-%d-00')


                    try:
                        val += str(dft)
                    except:
                        pass
                if j.dtTime and j.field.type == 'datetime':
                    if sepIdx>0 and j.field.multiSeperator:
                        val += str(j.field.multiSeperator)
                    if j.field.dateFormat and len(j.field.dateFormat)>0:
                        try:
                            dft = datetime.datetime.strftime(j.dtTime, str(j.field.dateFormat)+' %H:%M:%S')
                        except:
                            pass
                    else:
                        dft = j.dtTime

                    if 'Month&Year' == j.dtMode:
                        dft = datetime.datetime.strftime(j.dtTime, '%m-00-%Y')
                    elif 'Only Year' == j.dtMode:
                        dft = '00-00-' + str(j.dtTime.year)
                    elif 'date&month' == j.dtMode:
                        if not j.dtTime is None:
                            dft = datetime.datetime.strftime(j.dtTime, '%m-%d-00')


                    val += str(dft)
                if j.txt and j.field.type == 'text':
                    if sepIdx>0 and j.field.multiSeperator:
                        val += str(j.field.multiSeperator)
                    try:
                        val += str(j.txt)
                    except:
                        val += str(j.txt.encode("utf-8"))

                if j.field.type == 'boolean':
                    if sepIdx>0 and j.field.multiSeperator:
                        val += str(j.field.multiSeperator)
                    val += str(j.bool)
                if j.email and j.field.type == 'email':
                    if sepIdx>0 and j.field.multiSeperator:
                        val += str(j.field.multiSeperator)
                    val += str(j.email)
                if j.number and j.field.type == 'number':
                    if sepIdx>0 and j.field.multiSeperator:
                        val += str(j.field.multiSeperator)
                    val += str(j.number)
                if j.field.type == 'name':
                    name = ''
                    if j.prefix:
                        prefix = cleanStr(j.prefix).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                    else:
                        prefix = ''
                    if j.fname:
                        fname = cleanStr(j.fname).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                    else:
                        fname = ''
                    if j.mName:
                        mname = cleanStr(j.mName).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                    else:
                        mname = ''
                    if j.lname:
                        lname = cleanStr(j.lname).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                    else:
                        lname = ''
                    if j.compName:
                        compname = cleanStr(j.compName).replace('PN' , '__').replace('FN' , '**').replace('MN' , '^^').replace('LN' , '@@').replace('CN' , '##')
                    else:
                        compname = ''

                    nstrtemp = "".join([fname , mname , lname , compname , prefix])
                    if len(nstrtemp) == 0:
                        continue
                    j.field.nameFormat = j.field.nameFormat.replace('\\s' , ' ')
                    name = j.field.nameFormat.replace('FN' , cleanStr(fname)).replace('MN' , cleanStr(mname)).replace('LN' , cleanStr(lname)).replace('PN' , cleanStr(prefix)).replace('|' , '').replace('CN' , cleanStr(compname)).strip()
                    if name == "":
                        continue

                    nfData = str(j.field.nameFormat).split('|')
                    nfData.append(',')
                    for repeat in [0,1,2,3]:
                        name = cleanName(name , nfData).strip()
                    name = ' '.join(name.split())

                    if nfData[7] in name:
                        cIndx = name.index(nfData[7])
                        if len(name)>2 and name[cIndx-2] != " " and name[cIndx-2] in nfData:
                            inxRoRemove = cIndx -2
                            if len(name) > inxRoRemove:
                                name = name[0 : inxRoRemove : ] + name[inxRoRemove + 1 : :]
                        cIndx = name.index(nfData[7])
                        if name[cIndx-1] == " ":
                            inxRoRemove = cIndx -1
                            if len(name) > inxRoRemove:
                                name = name[0 : inxRoRemove : ] + name[inxRoRemove + 1 : :]

                    name = name.replace( '__', 'PN').replace( '**', 'FN').replace( '^^', 'MN' ).replace( '@@', 'LN').replace( '##', 'CN' )

                    if j.field.multiSeperator is None:
                        j.field.multiSeperator = ";"

                    if val == ",":
                        continue

                    if val == "":
                        val = name
                    else:
                        val += j.field.multiSeperator +  name


            if hdWidth[ridx+2]<=len(val):
                hdWidth[ridx+2] = len(val)+10
            hd1.append(val)

        Sheet1.append(hd1)
    for idx,i in enumerate(hdWidth):
        Sheet1.column_dimensions[alphaChars[idx]].width = i

    response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx'%(projectObj.title)
    return response



class ProjectDownloadAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        return projectReportExcel(request, False)



def calUserStats(userObj,fileObjs, lddassainData):
    toRet = {}
    # userObj = User.objects.get(pk=int(userPk))

    # fileObjs = FileUpload.objects.filter(pk__in=filePkList)
    codingVal = fileObjs.filter(coded=True,coder=userObj).count()
    qcVal = fileObjs.filter(checked=True,qa=userObj).count()
    reCodingVal = fileObjs.filter(recoded=True,recoder=userObj).count()
    reQcVal = fileObjs.filter(rechecked=True,reqa=userObj).count()
    if codingVal>0:
        wpercent = int(round((fileObjs.filter(coded=True,coder=userObj,recoded=True).count() * 100) / codingVal))
    else:
        wpercent = 0

    lddcoderCount = lddassainData.filter(coded=True,coder=userObj).aggregate(ct=Sum('count'))
    lddcoderCount = lddcoderCount['ct'] if lddcoderCount['ct'] else 0
    lddqcCount = lddassainData.filter(checked=True,qc=userObj).aggregate(ct=Sum('count'))
    lddqcCount = lddqcCount['ct'] if lddqcCount['ct'] else 0
    toRet['lddcoderCount'] = lddcoderCount
    toRet['lddqcCount'] = lddqcCount
    toRet['wpercent'] = wpercent
    toRet['codingCount'] = codingVal
    toRet['recodingCount'] = reCodingVal
    toRet['qcCount'] = qcVal
    toRet['reqcCount'] =reQcVal
    return toRet


def getFieldValue(user,fvlist):
    toRet = {}
    coding = fvlist.filter(file__coder = user).count()
    recoding = fvlist.filter(file__recoder = user).count()
    qa = fvlist.filter(file__qa = user).count()
    reqa = fvlist.filter(file__reqa = user).count()





    # for p in project.objects.all():
    #     coding += fvlist.filter(file__in = p.projectfiles.filter(coder = user)).count()*p.factor
    #     qa += fvlist.filter(file__in = p.projectfiles.filter(qa = user)).count()*p.factor
    #     recoding += fvlist.filter(file__in = p.projectfiles.filter(recoder = user)).count()*p.factor
    #     reqa += fvlist.filter(file__in = p.projectfiles.filter(reqa = user)).count()*p.factor

    total = coding + recoding + qa + reqa
    toRet['coding'] = coding
    toRet['recoding'] = recoding
    toRet['qa'] = qa
    toRet['reqa'] = reqa
    toRet['total'] = total
    # print total
    return toRet


class UserReportsAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        toRet = {}
        # print request.GET
        userObj = User.objects.get(pk=int(request.GET['userPk']))
        typ = request.GET['typ']
        if int(request.GET['project']) == 0:
            projectsPkList = list(project.objects.all().values_list('pk',flat=True))
        else:
            projectsPkList = list(project.objects.filter(pk=int(request.GET['project'])).values_list('pk',flat=True))
        # print projectsPkList,'projectsPkListprojectsPkList'
        stDate = datetime.datetime.strptime(str(request.GET['stDate']), '%Y-%m-%dT%H:%M:%S.%fZ')
        edDate = datetime.datetime.strptime(str(request.GET['edDate']), '%Y-%m-%dT%H:%M:%S.%fZ')
        if typ=='days':
            stDate = stDate.date()
            edDate = edDate.date()+relativedelta(days=1)
            filePkList = list(FieldValue.objects.filter(created__range=[stDate,edDate]).values_list('file',flat=True).distinct())
            lddassainData = LDDAssignment.objects.filter(created__range=[stDate,edDate],project__in=projectsPkList)
        else:
            filePkList = list(FieldValue.objects.filter(created__gte = stDate,created__lte = edDate).values_list('file',flat=True).distinct())
            lddassainData = LDDAssignment.objects.filter(created__gte = stDate,created__lte = edDate,project__in=projectsPkList)
        # print len(filePkList),'filePkListfilePkList',stDate,edDate

        fileObjs = FileUpload.objects.filter(pk__in=filePkList,project__in=projectsPkList)
        codingVal = fileObjs.filter(coded=True,coder=userObj).count()
        qcVal = fileObjs.filter(checked=True,qa=userObj).count()
        reCodingVal = fileObjs.filter(recoded=True,recoder=userObj).count()
        reQcVal = fileObjs.filter(rechecked=True,reqa=userObj).count()
        if codingVal>0:
            wpercent = int(round((fileObjs.filter(coded=True,coder=userObj,recoded=True).count() * 100) / codingVal))
        else:
            wpercent = 0

        lddcoderCount = lddassainData.filter(coded=True,coder=userObj).aggregate(ct=Sum('count'))
        lddcoderCount = lddcoderCount['ct'] if lddcoderCount['ct'] else 0
        lddqcCount = lddassainData.filter(checked=True,qc=userObj).aggregate(ct=Sum('count'))
        lddqcCount = lddqcCount['ct'] if lddqcCount['ct'] else 0
        toRet['lddcoderCount'] = lddcoderCount
        toRet['lddqcCount'] = lddqcCount
        toRet['wpercent'] = wpercent
        toRet['codingCount'] = codingVal + reCodingVal
        toRet['qcCount'] = qcVal + reQcVal
        return Response(toRet,status=status.HTTP_200_OK)

class GetOnlineReportsAPI(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def get(self, request, format=None):

        toRetVal = []
        if 'projPk' in request.GET:
            projectsData = project.objects.filter(pk=int(request.GET['projPk'])).order_by('-created')
        elif 'page' in request.GET:
            projectsData = project.objects.filter(~Q(status='complete')).order_by('-created')[int( request.GET['page'] )*10  : (int( request.GET['page'] ) + 1)*10]
        else:
            todays =  FieldValue.objects.filter(created__gt = datetime.datetime.today() - datetime.timedelta(days = 1)  ).values_list('project' , flat= True).distinct()
            projectsData = project.objects.filter(~Q(status='complete') , pk__in = todays).order_by('-created')
        # print len(projectsData),'projectsssssssss'
        for i in projectsData:
            statsData = projectStats(i)
            statsData['project'] = i.title
            statsData['status'] = i.status
            statsData['created'] = i.created
            statsData['pk'] = i.pk
            statsData['totalFiles'] = FileUpload.objects.filter(parent__isnull=True,project=i).count()
            fileObjs = FileUpload.objects.filter(parent__isnull=True,project=i)
            codersList = list(fileObjs.filter(coder__isnull=False).values_list('coder',flat=True).distinct())
            qaList = list(fileObjs.filter(qa__isnull=False).values_list('qa',flat=True).distinct())
            recodersList = list(fileObjs.filter(recoder__isnull=False).values_list('recoder',flat=True).distinct())
            reqcList = list(fileObjs.filter(reqa__isnull=False).values_list('reqa',flat=True).distinct())
            userStats = []

            if 'typ' in request.GET and request.GET['typ'] == 'all':
                print 'all report downloadddd'
            else:
                # print codersList,qaList,recodersList,reqcList
                usersPklist  = list(set(codersList+qaList+recodersList+reqcList))
                # print usersPklist
                for userId in usersPklist:
                    toRet={}
                    codingVal = fileObjs.filter(coded=True,coder__id=userId).count()
                    qcVal = fileObjs.filter(checked=True,qa__id=userId).count()
                    reCodingVal = fileObjs.filter(recoded=True,recoder__id=userId).count()
                    reQcVal = fileObjs.filter(rechecked=True,reqa__id=userId).count()
                    if codingVal>0:
                        wcount = fileObjs.filter(coded=True,coder__id=userId,recoded=True).count()
                        wpercent = int(round((wcount * 100) / codingVal))
                    else:
                        wpercent = 0
                        wcount = 0

                    lddcoderCount = LDDAssignment.objects.filter(project=i,coded=True,coder__id=userId).aggregate(ct=Sum('count'))
                    lddcoderCount = lddcoderCount['ct'] if lddcoderCount['ct'] else 0
                    lddqcCount = LDDAssignment.objects.filter(project=i,checked=True,qc__id=userId).aggregate(ct=Sum('count'))
                    lddqcCount = lddqcCount['ct'] if lddqcCount['ct'] else 0
                    toRet['lddcoderCount'] = lddcoderCount
                    toRet['lddqcCount'] = lddqcCount
                    toRet['wpercent'] = wpercent
                    toRet['wcount'] = wcount
                    toRet['codingCount'] = codingVal + reCodingVal
                    toRet['qcCount'] = qcVal + reQcVal
                    toRet['userpk'] = userId
                    userStats.append(toRet)

            statsData['userStats'] = userStats
            toRetVal.append(statsData)

        # print len(toRetVal),'objects returnnnnn'
        if request.GET.get('typ'):
            workbook = Workbook()
            Sheet1 = workbook.active
            hdFont = Font(size=12,bold=True)
            blue_font = Font(color='0029e6da', size=13,bold=True)
            alphaChars = list(string.ascii_uppercase)
            if request.GET['typ'] == 'all':
                Sheet1.title = 'Online Projects Report'

                hd = ['Project','Created','Status','Total Files','Coding','Qc','Re Coding','Re Qc']
                Sheet1.append(hd)
                for idx,i in enumerate(hd):
                    cl = str(alphaChars[idx])+'1'
                    Sheet1[cl].font = hdFont
                    if idx == 0:
                        Sheet1.column_dimensions[alphaChars[idx]].width = 30
                    else:
                        Sheet1.column_dimensions[alphaChars[idx]].width = 20

                for i in toRetVal:
                    hd1 = [i['project'],i['created'].date(),i['status'],i['totalFiles'],str(i['codingval'])+' ({0}%)'.format(str(i['coding'])),str(i['qcval'])+' ({0}%)'.format(str(i['qc'])),str(i['reCodingval'])+' ({0}%)'.format(str(i['recoding'])),str(i['reQcval'])+' ({0}%)'.format(str(i['reqc']))]

                    Sheet1.append(hd1)
            else:
                Sheet1.title = 'Project Report'

                hdp = ['Project','Created','Status','Total Files','Coding','Qc','Re Coding','Re Qc']
                Sheet1.append(hdp)

                for i in toRetVal:
                    hd1 = [i['project'],i['created'].date(),i['status'],i['totalFiles'],str(i['codingval'])+' ({0}%)'.format(str(i['coding'])),str(i['qcval'])+' ({0}%)'.format(str(i['qc'])),str(i['reCodingval'])+' ({0}%)'.format(str(i['recoding'])),str(i['reQcval'])+' ({0}%)'.format(str(i['reqc']))]

                    Sheet1.append(hd1)
                Sheet1.append([])
                Sheet1.append([])
                Sheet1.append([])

                hd = ['User Name','Total Coading','Total Qc','Total LDD Coding','Total LDD Qc','Wrong']
                Sheet1.append(hd)
                for idx,i in enumerate(hdp):
                    cl = str(alphaChars[idx])+'1'
                    cl6 = str(alphaChars[idx])+'6'
                    Sheet1[cl].font = blue_font
                    Sheet1[cl6].font = hdFont
                    Sheet1.column_dimensions[alphaChars[idx]].width = 30

                for i in toRetVal[0]['userStats']:
                    userObj = User.objects.get(pk=int(i['userpk']))
                    # print userObj
                    if userObj.first_name:
                        name = userObj.first_name
                        if userObj.last_name:
                            name += ' '+userObj.last_name
                    else:
                        name = ''
                    hd1 = [name,i['codingCount'],i['qcCount'],i['lddcoderCount'],i['qcCount'],str(i['wcount'])+' ( {0}% )'.format(str(i['wpercent']))]

                    Sheet1.append(hd1)


            # print 'return excellllll'
            response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=onlineReport.xlsx'
            return response
        else:
            return Response(toRetVal,status=status.HTTP_200_OK)

class GetProjectReportsAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        toRet = {'codingData':[],'codingLabels':[],'qcData':[],'qcLabels':[],'recodingData':[],'recodingLabels':[],'reqcData':[],'reqcLabels':[]}
        # print request.GET
        if 'projectPk' in request.GET and 'stDate' in request.GET and 'edDate' in request.GET:
            stDate = datetime.datetime.strptime(str(request.GET['stDate']), '%Y-%m-%dT%H:%M:%S.%fZ')
            edDate = datetime.datetime.strptime(str(request.GET['edDate']), '%Y-%m-%dT%H:%M:%S.%fZ')
            projectsPkList = projectsPkList = list(project.objects.filter(pk=int(request.GET['projectPk'])).values_list('pk',flat=True))
            duration_insec = (edDate-stDate).total_seconds()
            durInHrs = int(divmod(duration_insec, 3600)[0])
            # print durInHrs,'durInHrsdurInHrs'
            for i in range(durInHrs):
                st = stDate
                ed = st + relativedelta(hours=1)
                filePkList = list(FieldValue.objects.filter(created__gte = st,created__lte = ed).values_list('file',flat=True).distinct())
                fileObjs = FileUpload.objects.filter(pk__in=filePkList,project__in=projectsPkList,parent__isnull=True)
                codingVal = fileObjs.filter(coded=True).count()
                qcVal = fileObjs.filter(checked=True).count()
                reCodingVal = fileObjs.filter(recoded=True).count()
                reQcVal = fileObjs.filter(rechecked=True).count()
                toRet['codingData'].append(codingVal)
                toRet['qcData'].append(qcVal)
                toRet['recodingData'].append(reCodingVal)
                toRet['reqcData'].append(reQcVal)
                chSt = st+relativedelta(hours=5,minutes=30)
                ched = ed+relativedelta(hours=5,minutes=30)
                toRet['codingLabels'].append(str(chSt.hour)+'-'+str(ched.hour))
                toRet['qcLabels'].append(str(chSt.hour)+'-'+str(ched.hour))
                toRet['recodingLabels'].append(str(chSt.hour)+'-'+str(ched.hour))
                toRet['reqcLabels'].append(str(chSt.hour)+'-'+str(ched.hour))
                stDate = ed
        return Response(toRet,status=status.HTTP_200_OK)

def getPerformanceStats(startDate, endDate, request):
    fvl =  FieldValue.objects.filter(created__range=(startDate, endDate))
    lddassainData = LDDAssignment.objects.filter(updated__gte = startDate, updated__lte=endDate)
    fileList = FileUpload.objects.filter(lastupdated__gte = startDate , lastupdated__lte=endDate )
    # filePkList = list(FieldValue.objects.filter(updated__gte = startDate , updated__lte=endDate ).values_list('file',flat=True).distinct())

    userStatsData = calUserStats(request.user,fileList , lddassainData)
    userFieldValue = getFieldValue(request.user,fvl)
    try:
        trgt = request.user.designation.salaryGrade.target
    except:
        trgt = 'NA'

    usertaskObj = task.objects.filter(dueDate__gte=startDate,dueDate__lte=endDate, to = request.user)
    try:
        totalpoints = 0
        for j in usertaskObj:
            points = 0
            points =  (float(j.conversionFactor)*(float(j.completion)/100.0))
            totalpoints+= points
    except:
        totalpoints=0

    return {"counts" : userStatsData , "points" : userFieldValue , "target" : trgt , "generalTasks" : totalpoints  }

class GetPerformanceReportAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        startDate = request.GET['stDate']
        endDate = request.GET['edDate']
        return Response(getPerformanceStats(startDate , endDate , request),status=status.HTTP_200_OK)

from django.core.paginator import Paginator

class DownloadPerformanceReportAPI(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        startDate = request.GET['stDate']
        endDate = request.GET['edDate']
        workbook = Workbook()
        Sheet1 = workbook.active
        Sheet1.title = 'Performance Report'
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        projectObj = None
        hd = ['Name','User ID','Total Coding','Total Qc','Total LDD Coding','Total LDD Qc','Wrong Percentage (%)','Task Points','Coding','Recoding','QA','Re-QA','Total', 'Target']

        Sheet1.append(hd)
        hdWidth = []
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            if idx == 0 or idx == len(hd)-1:
                hdWidth.append(25)
            else:
                hdWidth.append(20)

        # projectCodersTeam = list(project.objects.filter(coderTeam__isnull=False , created__gt = startDate ).values_list('coderTeam',flat=True))
        # projectQcTeam = list(project.objects.filter(qaTeam__isnull=False, created__gt = startDate).values_list('qaTeam',flat=True))
        # projectReCodersTeam = list(project.objects.filter(reCoderTeam__isnull=False, created__gt = startDate).values_list('reCoderTeam',flat=True))
        # projectReQcTeam = list(project.objects.filter(reQaTeam__isnull=False, created__gt = startDate).values_list('reQaTeam',flat=True))

        # teamsList = list(set(projectCodersTeam + projectQcTeam + projectReCodersTeam + projectReQcTeam))
        # print teamsList
        # userdesigObjsAll = Paginator(, 80)
        # print userdesigObjs

        userdesigObjs = designation.objects.filter(team = request.GET['team'] ).distinct()

        today = datetime.datetime.now()
        taskObj = task.objects.filter(dueDate__gte=startDate,dueDate__lte=endDate)

        fvlist = FieldValue.objects.filter(created__range=(startDate, endDate))

        lddassainData = LDDAssignment.objects.filter(created__gt = startDate, created__lte=endDate)
        # filePkList = list(FieldValue.objects.filter(created__gte = startDate , created__lte=endDate ).values_list('file',flat=True).distinct())
        fileList = FileUpload.objects.filter(lastupdated__gte = startDate , lastupdated__lte=endDate )
        for i in userdesigObjs:
            hd1 = []
            dt = (today - relativedelta(days=30)).date()

            userStatsData = calUserStats(i.user,fileList,lddassainData)
            userFieldValue = getFieldValue(i.user, fvlist)
            try:
                usertaskObj = taskObj.filter(to = i.user)
                totalpoints = 0
                for j in usertaskObj:
                    points = 0
                    points =  (float(j.conversionFactor)*(float(j.completion)/100.0))
                    totalpoints+= points
            except:
                totalpoints=0
            name = ''
            if i.user.first_name:
                name += i.user.first_name
            if i.user.last_name:
                name = name+' '+i.user.last_name

            hd1.append(name)
            hd1.append('# '+str(i.user.pk))
            hd1.append(userStatsData['codingCount'])
            hd1.append(userStatsData['qcCount'])
            hd1.append(userStatsData['lddcoderCount'])
            hd1.append(userStatsData['lddqcCount'])
            hd1.append(str(userStatsData['wpercent']) + '%')
            hd1.append(totalpoints)
            hd1.append(userFieldValue['coding'])
            hd1.append(userFieldValue['recoding'])
            hd1.append(userFieldValue['qa'])
            hd1.append(userFieldValue['reqa'])
            hd1.append(userFieldValue['total'])
            try:
                hd1.append( i.salaryGrade.target )
            except:
                hd1.append('NA')
            # for widx,w in enumerate(hd1):
            #     if hdWidth[widx]<=len(str(hd1[widx])):
            #         hdWidth[widx] = len(str(hd1[widx]))+10

            Sheet1.append(hd1)

        for idx,i in enumerate(hdWidth):
            Sheet1.column_dimensions[alphaChars[idx]].width = i

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        title = "Performance report"
        if projectObj:
            title = projectObj.title

        response['Content-Disposition'] = 'attachment; filename=%s.xlsx'%(title)
        return response

class AbbyImportDataAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        toRet = {}
        projectObj = project.objects.get(pk=int(request.data['projectPk']))
        projectFileObjs = FileUpload.objects.filter(project=projectObj,parent__isnull=True)
        projectFilesNames = []
        for f in projectFileObjs:
            if f.path:
                fPathVal = str(f.path).split('/')[-1]
            else:
                fPathVal = str(f.file).split('/')[-1]
                if '__' in fPathVal:
                    fPathVal = fPathVal.split('__')[-1]
            projectFilesNames.append(fPathVal)
        # print projectFilesNames

        df = pd.read_excel(request.FILES['dataFile'])
        lstDic = df.to_json(orient='split')
        lstDic = json.loads(lstDic)

        fields = lstDic['columns']
        fieldObjs = []
        for f in fields[2:]:
            fieldObjs.append(ProjectFields.objects.filter(project=projectObj,name=str(f)).first())
        # print fieldObjs
        fieldsData = lstDic['data']
        fieldValuesData = []
        docsPkList = []
        for rIdx,i in enumerate(fieldsData):
            rowData = i[2:]
            # print rIdx
            if any(rowData):
                for clIdx,cl in enumerate(rowData):
                    if cl is not None and fieldObjs[clIdx].type!='name' and not fieldObjs[clIdx].multi:
                        # print clIdx
                        print rIdx,projectFilesNames[rIdx],cl,'cllllllllllll',clIdx,fieldObjs[clIdx].name,fieldObjs[clIdx].type

                        data = {'project':projectObj,'field':fieldObjs[clIdx],'file':projectFileObjs[rIdx]}
                        if fieldObjs[clIdx].type=='datetime':
                            dt = datetime.datetime.strptime(str(cl), str(fieldObjs[clIdx].dateFormat)+' %H:%M:%S')
                            data['dtTime']=dt
                        elif fieldObjs[clIdx].type=='date':
                            dt = datetime.datetime.strptime(str(cl), str(fieldObjs[clIdx].dateFormat))
                            data['dt']=dt
                        elif fieldObjs[clIdx].type=='email':
                            data['email']=str(cl)
                        elif fieldObjs[clIdx].type=='text':
                            data['txt']=str(cl)
                        elif fieldObjs[clIdx].type=='number':
                            data['number']=str(cl)
                        elif fieldObjs[clIdx].type=='smallText':
                            data['txt']=str(cl)
                        elif fieldObjs[clIdx].type=='boolean':
                            data['bool']=cl
                        else:
                            continue
                        print "Came here"
                        fieldValuesData.append(FieldValue(**data))
                        print "Came here 2"
                        docsPkList.append(projectFileObjs[rIdx].pk)
        if len(fieldValuesData)>0:
            FieldValue.objects.bulk_create(fieldValuesData)
            FileUpload.objects.filter(pk__in=docsPkList).update(status='coded',coded=True)
        print len(fieldValuesData),'fieldValuesData has been createdddddd'

        toRet['createdRecords'] = len(fieldValuesData)
        return Response(toRet,status=status.HTTP_200_OK)

from django.template import Template, Context
projectStartTemplate = Template(open( os.path.join(globalSettings.BASE_DIR , 'projects', 'templates','project.html')).read())
projectRecoderTemplate = Template(open( os.path.join(globalSettings.BASE_DIR , 'projects', 'templates','projectRecoder.html')).read())
projectQATemplate = Template(open( os.path.join(globalSettings.BASE_DIR , 'projects', 'templates','projectQa.html')).read())
projectreQATemplate = Template(open( os.path.join(globalSettings.BASE_DIR , 'projects', 'templates','projectreQa.html')).read())


@login_required(login_url = globalSettings.LOGIN_URL)
def projeCtcoderStart(request,pid=None ):
    if pid:
        data = {'typ':'coder','imageViewerUrl':globalSettings.IMAGEVIEWER_URL}
        data['projectid'] = pid
        data['projectname'] = "Coding : " + project.objects.get(pk = pid).title
        data['projectnameBase'] = project.objects.get(pk = pid).title

        u = request.user

        userTams = list(u.designation.team.all().values_list('pk',flat=True))
        toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) | Q(coderTeam__in = userTams) | Q(reCoderTeam__in = userTams))
        projs = toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')

        print "Proj count " , projs.filter(pk = pid).count()

        if projs.filter(pk = pid).count()==0:
            return redirect('lightHome')

        return HttpResponse(projectStartTemplate.render(Context(data)))
        # return render(request, 'project.html' , data)
    else:
       print 'elseeeeeeeeeeee'
       return HttpResponse('', status=status.HTTP_404_NOT_FOUND)


@login_required(login_url = globalSettings.LOGIN_URL)
def projectQaStart(request ,pid=None):
    if pid:
        data = {'typ':'qa','imageViewerUrl':globalSettings.IMAGEVIEWER_URL}
        data['projectid'] = pid
        data['projectname'] = "QC : " + project.objects.get(pk = pid).title
        u = request.user

        userTams = list(u.designation.team.all().values_list('pk',flat=True))
        toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) |  Q(qaTeam__in = userTams) | Q(reQaTeam__in = userTams))
        projs = toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')


        if projs.filter(pk = pid).count()==0:
            return redirect('lightHome')

        data['canEdit'] = project.objects.get(pk = int(pid)).qaCanEdit
        return HttpResponse(projectQATemplate.render(Context(data)))
        # return render(request, 'projectQa.html' , data)
    else:
       print 'elseeeeeeeeeeee'
       return HttpResponse('', status=status.HTTP_404_NOT_FOUND)

@login_required(login_url = globalSettings.LOGIN_URL)
def projectRecoderStart(request ,pid=None):
    if pid:
        data = {'typ':'recoding','imageViewerUrl':globalSettings.IMAGEVIEWER_URL}
        data['projectid'] = pid
        data['projectname'] = "RECODING : " + project.objects.get(pk = pid).title

        u = request.user

        userTams = list(u.designation.team.all().values_list('pk',flat=True))
        toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) | Q(coderTeam__in = userTams) | Q(reCoderTeam__in = userTams))
        projs = toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')


        if projs.filter(pk = pid).count()==0:
            return redirect('lightHome')

        data['canEdit'] = project.objects.get(pk = int(pid)).qaCanEdit
        return HttpResponse(projectRecoderTemplate.render(Context(data)))
        # return render(request, 'projectRecoder.html' , data)
    else:
       print 'elseeeeeeeeeeee'
       return HttpResponse('', status=status.HTTP_404_NOT_FOUND)

@login_required(login_url = globalSettings.LOGIN_URL)
def projectreCCStart(request ,pid=None, fid = None):
    if pid:
        data = {'typ':'CC','imageViewerUrl':globalSettings.IMAGEVIEWER_URL}
        data['projectid'] = pid
        data['fileid'] = fid
        data['projectname'] = "CC : " + project.objects.get(pk = pid).title


        print "file ID : " , fid
        u = request.user

        userTams = list(u.designation.team.all().values_list('pk',flat=True))
        toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) | Q(coderTeam__in = userTams) | Q(reCoderTeam__in = userTams))
        projs = toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')


        if projs.filter(pk = pid).count()==0:
            return redirect('lightHome')

        data['canEdit'] = project.objects.get(pk = int(pid)).qaCanEdit
        return render(request, 'projectCC.html' , data)
    else:
       print 'elseeeeeeeeeeee'
       return HttpResponse('', status=status.HTTP_404_NOT_FOUND)


@login_required(login_url = globalSettings.LOGIN_URL)
def projectreQaStart(request,pid=None):
    if pid:
        data = {'typ':'reqa','imageViewerUrl':globalSettings.IMAGEVIEWER_URL}
        data['projectid'] = pid
        data['projectname'] = "RE-QC : " + project.objects.get(pk = pid).title
        u = request.user

        userTams = list(u.designation.team.all().values_list('pk',flat=True))
        toRet = u.projectsInvolvedIn.all() | project.objects.filter(Q(user = u) |  Q(qaTeam__in = userTams) | Q(reQaTeam__in = userTams))
        projs = toRet.filter(~Q(status='complete'),disabled=False).distinct().order_by('-created')

        if projs.filter(pk = pid).count()==0:
            return redirect('lightHome')

        data['canEdit'] = project.objects.get(pk = int(pid)).qaCanEdit
        return HttpResponse(projectreQATemplate.render(Context(data)))
        return render(request, 'projectreQa.html' , data)
    else:
       print 'elseeeeeeeeeeee'
       return HttpResponse('', status=status.HTTP_404_NOT_FOUND)

class UpdateLDDAssignmentAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        toRet = {}
        postData = request.data
        print postData['pk'],postData
        lddObj = LDDAssignment.objects.get(pk=int(str(postData['pk'])))
        if 'cmt' in postData:
            print str(postData['cmt'])
            lddObj.comment = str(postData['cmt'])
            lddObj.checked = True
            lddObj.qcOutput = json.dumps(json.loads(postData['files']))
            lddObj.pagesCount = int(postData['count'])
            lddObj.save()
        elif 'codercmt' in postData:
            print str(postData['codercmt'])
            lddObj.coderComment = str(postData['codercmt'])
            lddObj.coded = True
            lddObj.pagesCount = int(postData['count'])
            lddObj.output = json.dumps(json.loads(postData['files']))
            lddObj.save()

        return Response(toRet,status=status.HTTP_200_OK)

class ProjectCompletionCheckAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        toRet = {}
        projectObj = project.objects.get(pk = request.GET['projectPk'])
        flObj = FileUpload.objects.filter(project_id=projectObj)
        if flObj.filter(parent__isnull=True).count() == flObj.filter(status='approved').count():
            projectObj.status = 'complete'
            toRet['projectStatus'] = 'completed'
            projectObj.save()
        return Response(toRet,status=status.HTTP_200_OK)

class UploadLDDFilesAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        lines = request.FILES['file'].read()
        folders = []
        files = []
        for line in lines.split('\n'):
            fileName = line.split('\\')[-1]
            folder = "\\".join(line.split('\\')[0:-1])
            print folder , fileName
            if len(folder.strip()) == 0:
                continue
            files.append({"folder" : folder  , "fileName"  : fileName })
            if folder not in folders:
                folders.append(folder)
        folderCountMap = []
        for fl in folders:
            count = 0
            filesList = []
            for fil in files:
                if fil['folder'] == fl:
                    count +=1
                    filesList.append(fil["fileName"])

            folderCountMap.append({"folder" : fl , "count" : count , "files" : ",".join(filesList) , "coded" : False , "checked" : False})

        return Response(folderCountMap,status=status.HTTP_200_OK)

class AssignLDDAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):

        p = project.objects.get(pk = request.GET['project'])

        for assignment in request.data['assignments']:
            print assignment
            lddAss = LDDAssignment(project = p , coder = User.objects.get(pk = assignment['coder']['pk']) , qa  = User.objects.get(pk = assignment['coder']['pk']) , folder = assignment['folder'] , files = assignment['files'] , count = assignment['count'])

            lddAss.save()

        return Response({},status=status.HTTP_200_OK)

import csv
from excel_response import ExcelResponse
class LDDDownloadAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
            LDDOutout = {}
            fileLists = []
            for ldd in LDDAssignment.objects.filter(project__id = request.GET['projectpk']).order_by('pk'):
                # print "Using : " , ldd.pk
                if ldd.qcOutput is not None:
                    coder = ldd.coder.first_name + ' ' + ldd.coder.last_name
                    qc = ldd.qc.first_name + ' '+ ldd.qc.last_name
                    fileListsTemp = json.loads(str(ldd.qcOutput))["files"]
                    for flt in fileListsTemp:
                        print flt
                        fileLists.append(flt)




            for fli, fl in enumerate(fileLists):
                fileLists[fli]['count'] = 1

            for fli, fl in enumerate(fileLists):
                if fl['documentGroup'] is not None:
                    for fl1i , fl1 in enumerate(fileLists):
                        if fl1['id'] == fl['documentGroup']:
                            fileLists[fl1i]['count'] += 1

            # print "file list", fileLists
            # print "LDDOutout" , LDDOutout
            for fl in fileLists:
                # print fl
                if fl['documentGroup'] is not None:
                    print "key : " , fileLists[fl['documentGroup']]["name"]
                    LDDOutout[fileLists[fl['documentGroup']]["name"]]["end"] = fl["name"]
                    LDDOutout[fileLists[fl['documentGroup']]["name"]]["coder"] = coder
                    LDDOutout[fileLists[fl['documentGroup']]["name"]]["qc"] = qc
                    LDDOutout[fileLists[fl['documentGroup']]["name"]]["comment"] = ldd.comment
                    LDDOutout[fileLists[fl['documentGroup']]["name"]]["folder"] = fl["lddDir"]
                else:
                    LDDOutout[fl["name"]] = {"end":None , "astart": None , "aend":None , "id":fl["id"] , "count": fl["count"], 'coder': coder , 'qc': qc , 'comment' : ldd.comment,"folder":fl["lddDir"]}
                if fl['attachmentGroup'] is not None:
                    LDDOutout[fileLists[fl['attachmentGroup']]["name"]]["aend"] = fl["name"]
                    LDDOutout[fileLists[fl['attachmentGroup']]["name"]]["astart"] = fileLists[fl['attachmentGroup']]["name"]
                    LDDOutout[fileLists[fl['attachmentGroup']]["name"]]["coder"] = coder
                    LDDOutout[fileLists[fl['attachmentGroup']]["name"]]["qc"] = qc
                    LDDOutout[fileLists[fl['attachmentGroup']]["name"]]["comment"] = ldd.comment
                    LDDOutout[fileLists[fl['attachmentGroup']]["name"]]["folder"] = fl["lddDir"]



            LDDFinalOut = []
            for key in LDDOutout:
                value = LDDOutout[key]
                if value["end"]  is None:
                    end = key
                else:
                    end = value["end"]
                LDDFinalOut.append({"start": key , "end" : end, "astart": value["astart"] , "aend": value["aend"] , "id" : value["id"], 'count': value['count'] , 'coder' : value['coder'] , 'qc': value['qc'] , 'comment' : value['comment'],'directory':value['folder'] })


            LDDSorted = sorted(LDDFinalOut, key = lambda i: i['id'])
            for lddi, ldd in enumerate(LDDSorted):
                if ldd['aend'] is not None:
                    for lddi2 , ldd2 in enumerate(LDDSorted[lddi:]):
                        LDDSorted[lddi2+lddi]['astart'] = ldd['astart']
                        LDDSorted[lddi2+lddi]['aend'] = ldd['aend']
                        if ldd2['start']==ldd['aend']:
                            break
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="somefilename.csv"'

            if request.GET['type'] == 'excel':
                excelldd = []
                excelldd.append(['BEG DOC', 'END DOCT', 'BEG ATTACH', 'END ATTACH', 'PG COUNT','Directory'])
                for ldd in LDDSorted:
                    excelldd.append([ldd['start'], ldd['end'], ldd['astart'], ldd['aend'] , ldd['count'],ldd['directory']])
                if 'saveExcel' in request.GET:
                    hed = excelldd.pop(0)
                    df = pd.DataFrame(excelldd,columns=hed)
                    responseFile = os.path.join(globalSettings.BASE_DIR, 'media_root', 'lddExcels','excel_{0}.xlsx'.format(int(request.GET['projectpk'])))
                    try:
                        from io import BytesIO as IO # for modern python
                    except ImportError:
                        from io import StringIO as IO # for legacy python
                    xlwriter = pd.ExcelWriter(responseFile, engine='xlsxwriter')
                    df.to_excel(xlwriter, 'Sheet 1',index=False)
                    xlwriter.save()
                    return Response({'savedExcel':'excel_{0}.xlsx'.format(int(request.GET['projectpk']))},status=status.HTTP_200_OK)
                return ExcelResponse(excelldd)
            elif request.GET['type'] == 'psv':
                writer = csv.writer(response)
                writer.writerow(['BEG DOC', 'END DOCT', 'BEG ATTACH', 'END ATTACH', 'PG COUNT','Directory'])
                for ldd in LDDSorted:
                    writer.writerow([ldd['start'], ldd['end'], ldd['astart'], ldd['aend'] , ldd['count'],ldd['directory']])
                return response

            elif request.GET['type'] == 'csv':

                writer = csv.writer(response)
                writer.writerow(['BEG DOC', 'END DOCT', 'BEG ATTACH', 'END ATTACH', 'PG COUNT','Directory'])
                for ldd in LDDSorted:
                    writer.writerow([ldd['start'], ldd['end'], ldd['astart'], ldd['aend'] , ldd['count'],ldd['directory']])
                return response

            elif request.GET['type'] == 'withcoder':

                excelldd = []
                excelldd.append(['BEG DOC', 'END DOCT', 'BEG ATTACH', 'END ATTACH', 'PG COUNT', 'Remarks','QC Name', 'Coder Name','Directory'])
                for ldd in LDDSorted:
                    print ldd
                    excelldd.append([ldd['start'], ldd['end'], ldd['astart'], ldd['aend'] , ldd['count'] , ldd['comment'] , ldd['qc'] , ldd['coder'],ldd['directory']])
                return ExcelResponse(excelldd)


class GetProjectStatusAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        prj = project.objects.get(pk = request.GET['projectpk'])
        toRet = projectStats(prj)

        return Response(toRet,status=status.HTTP_200_OK)

class ClearOldFileValuesAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated,)
    def get(self, request, format=None):
        filesObj = FieldValue.objects.filter(project__id=int(request.GET['projectpk']),file__id=int(request.GET['filepk']))
        deletedCount = filesObj.count()
        filesObj.delete()
        return Response({'status':'Done','deletedCount':deletedCount},status=status.HTTP_200_OK)


from django.db import connection
import mysql.connector

class UploadLDDFileAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):

        HOST = globalSettings.DATABASES['default']['HOST']
        USER = globalSettings.DATABASES['default']['USER']
        PASSWORD = globalSettings.DATABASES['default']['PASSWORD']
        NAME = globalSettings.DATABASES['default']['NAME']

        mydb = mysql.connector.connect(
          host= HOST ,
          user= USER,
          passwd= PASSWORD,
          database= NAME
        )



        direc = request.data['baseDir']
        prj = project.objects.get(pk = request.data['project'])
        FileUpload.objects.filter(project = prj).delete()
        # prj.uploadingBreakFile = True
        prj.lddUploadedOn = datetime.datetime.now()
        prj.save()

        if 'savedFile' in request.data:
            ldd = os.path.join(globalSettings.BASE_DIR, 'media_root', 'lddExcels',str(request.data['savedFile']))
        else:
            ldd = request.FILES['ldd']
        # print ldd , direc , prj
        df = pd.read_excel(ldd)
        # print "original : " , df
        df = df.replace('_' , '#', regex=True)
        # print df

        lstDic = df[df.columns[0]].tolist()
        noExtlstDic = [str(i).split('.')[0] for i in lstDic]
        secColumn = df[df.columns[1]].tolist()
        noExtsecColumn = [str(i).split('.')[0] for i in secColumn]
        countColumn = df[df.columns[4]].tolist()
        try:
            thdColumn = df[df.columns[2]].tolist()
        except:
            thdColumn = []
        try:
            dirColumn = df[df.columns[5]].tolist()
        except:
            dirColumn = []

        minFileNum = re.findall(r'\d+', lstDic[0])[0]
        maxFileNum = re.findall(r'\d+', secColumn[-1])[0]
        stIdx = str(lstDic[0]).find(minFileNum)
        # print stIdx
        flStatic = str(noExtlstDic[0]).replace(str(noExtlstDic[0])[stIdx:stIdx+len(minFileNum)],'__')
        filObjDic = {}
        parentObjDic = {}
        dirPath = None
        parentFile = None
        fObjArray = []
        # print sum(countColumn),int(maxFileNum)
        prentNameList = []
        if False and sum(countColumn) == int(maxFileNum):
            print 'old logiccc'
            for i in range(int(minFileNum),int(maxFileNum)+1):
                f = flStatic

                print "f======" , f

                strNum = str(i).zfill(len(minFileNum))
                flNam = f.replace('__',strNum)
                if flNam in noExtlstDic:
                    dIdx = noExtlstDic.index(flNam)
                    if len(dirColumn)>0:
                        dirPath = dirColumn[dIdx]
                    flNam += '.' + str(lstDic[dIdx]).split('.')[-1]
                    parentFile = flNam
                elif flNam in noExtsecColumn:
                    dIdx = noExtsecColumn.index(flNam)
                    if len(dirColumn)>0:
                        dirPath = dirColumn[dIdx]
                    flNam += '.' + str(secColumn[dIdx]).split('.')[-1]
                    parentObjDic[flNam] = parentFile
                else:
                    flNam += '.' + str(parentFile).split('.')[-1]
                    parentObjDic[flNam] = parentFile
                if dirPath:
                    fld = dirPath
                    if not str(fld).endswith('/'):
                        fld += '/'
                else:
                    fld = ''
                if str(direc).endswith('/'):
                    flPath = str(direc)+fld+flNam
                else:
                    flPath = str(direc)+'/'+fld+flNam
                fObj = FileUpload.objects.create(**{'project':prj,'serverPath':direc,'path':flPath.replace('#' , '_')})
                filObjDic[flNam] = fObj
        else:
            createQueries = []
            print 'new logiccc'




            for fIdx,i in enumerate(lstDic):
                minFileNum = re.findall(r'\d+', lstDic[fIdx])[-1]
                maxFileNum = re.findall(r'\d+', secColumn[fIdx])[-1]
                stIdx = str(lstDic[fIdx]).find(minFileNum)
                # print " min max file num : ", stIdx , minFileNum , maxFileNum
                flStatic = str(lstDic[fIdx]).replace(str(lstDic[fIdx])[stIdx:stIdx+len(minFileNum)],'__')
                dirPath = None
                parentFile = None
                for j in range(int(minFileNum),int(maxFileNum)+1):
                    DEFAULT_EXTENSION = False
                    f = flStatic
                    # print "f======" , f , j
                    strNum = str(j).zfill(len(minFileNum))
                    flNam = f.replace('__',strNum)
                    parentNameSet = False
                    if flNam in lstDic:
                        # print "if : " , j
                        dIdx = lstDic.index(flNam)
                        if len(dirColumn)>0:
                            dirPath = dirColumn[dIdx]
                        parentFile = flNam
                    elif flNam in secColumn:
                        # print "Elseif 2 : " , j
                        dIdx = secColumn.index(flNam)
                        if len(dirColumn)>0:
                            dirPath = dirColumn[dIdx]
                        parentObjDic[flNam] = parentFile
                        parentNameSet = True
                    else:
                        # print "Else : " , j
                        parentObjDic[flNam] = parentFile
                        DEFAULT_EXTENSION = True
                        parentNameSet = True

                    if dirPath:
                        fld = dirPath
                        if not str(fld).endswith('/'):
                            fld += '/'
                    else:
                        fld = ''

                    parentName = None
                    if str(direc).endswith('/'):
                        flPath = str(direc)+fld+flNam
                        if parentNameSet:
                            parentName = parentFile.replace('#' , '_')
                    else:
                        flPath = str(direc)+'/'+fld+flNam
                        if parentNameSet:
                            parentName = parentFile.replace('#' , '_')
                    # print "creating : " , {'project':prj,'serverPath':direc,'path':flPath.replace('#' , '_')}
                    if DEFAULT_EXTENSION:
                        flPath = '.'.join( flPath.split('.')[:-1] ) + '.tif'
                        # print " flPath" , flPath

                    # query = "INSERT INTO `projects_fileupload` (`id`, `created`, `path`, `order`, `project_id`, `file`, `status`, `user_id`, `checked`, `coded`, `coder_id`, `locked`, `mistake`, `parent_id`, `qa_id`, `rechecked`, `recoded`, `recoder_id`, `reqa_id`, `serverPath`, `skiped`, `checklocked`, `rechecklocked`, `relocked`, `attachment_id`, `codeTime`, `qcTime`, `reCodeTime`, `reQcTime`, `coderComment`, `qcComment`, `recoderComment`, `reqcComment`, `lastupdated`) VALUES (NULL, CURRENT_TIMESTAMP, '%s', NULL, '%s', '', 'created', NULL, '0', '0', NULL, '0', '0', NULL, NULL, '0', '0', NULL, NULL, '%s', '0', '0', '0', '0', NULL, '0', '0', '0', '0', NULL, NULL, NULL, NULL, CURRENT_TIMESTAMP)"%(flPath.replace('#' , '_') , prj.pk, direc )

                    # createQueries.append(query)
                    if parentName is not None and parentName not in prentNameList:

                        prentNameList.append(parentName)

                    # fObj = FileUpload.objects.create(**{'project':prj,'serverPath':direc,'path':flPath.replace('#' , '_')})
                    if parentName == flPath.replace('#' , '_'):
                        fObj = FileUpload(**{'project':prj,'serverPath':direc,'path':flPath.replace('#' , '_') })
                    else:
                        fObj = FileUpload(**{'project':prj,'serverPath':direc,'path':flPath.replace('#' , '_') , 'parentName' : parentName})

                    fObjArray.append(fObj)
                    filObjDic[flNam] = fObj
        FileUpload.objects.bulk_create(fObjArray)
        print "Updating the parent child relationships"
        with connection.cursor() as cursor:
            # f = open('insert.sql' , 'w')
            # f.write(";\n".join(createQueries))
            # f.close()
            updateQueries = []
            newCursor = mydb.cursor()
            for pn in prentNameList:

                    query = "UPDATE projects_fileupload as parentTable SET parentTable.parent_id= (SELECT * from (SELECT id from projects_fileupload AS tempfiles where path like '%{0}' and project_id={1}) AS tf)  where parentName like '%{2}' and project_id={3}".format(pn , prj.pk , pn , prj.pk)
                    # updateQueries.append(query)
                    cursor.execute(query)

                    # newCursor.execute(query)

                    #UPDATE `projects_fileupload` as parentTable SET parentTable.parent_id= (SELECT * from (SELECT id from projects_fileupload AS tempfiles where path like '%00002794.tif' and project_id=1) AS tf)  where path like '%00002793.tif' and project_id=1

                    # upObj =  FileUpload.objects.filter(path__endswith =filObjDic[i].path ,project=prj ).update(parent = FileUpload.objects.get(path__endswith =parentObjDic[i].replace('#' , '_') ,project=prj ))
            # f = open('update.sql' , 'w')
            # f.write(";\n".join(updateQueries))
            # f.close()

            if len(thdColumn)>0:
                for idx,i in enumerate(lstDic):
                    if type(thdColumn[idx]) != float and lstDic[idx] != thdColumn[idx]:
                        FileUpload.objects.filter(path__endswith =filObjDic[i].path ,project=prj ).update(attachment = FileUpload.objects.get(path__endswith =thdColumn[idx].replace('#' , '_') ,project=prj ))

        print "renaming"
        try:
            wb = openpyxl.load_workbook(ldd)
            worksheet = wb["FILES"]
            for row in worksheet.iter_rows():
                # print row[0].value
                filName = row[0].value.split('.')[0]
                ext = row[0].value.split('.')[1]
                if ext == 'tif':
                    # print "Continuing"
                    continue
                # print filName , ext
                # for fu in FileUpload.objects.filter(path__contains = filName , project = prj ):
                #     print fu.project.title
                #     print fu.path
                #     print fu.pk
                #     print '----------'

                try:
                    fu = FileUpload.objects.get(path__contains = filName , project = prj )
                    fu.path = fu.path.split('.')[0] + '.' + ext
                    fu.save()
                except:
                    pass
        except:
            traceback.print_exc(file=sys.stdout)

        prjT = project.objects.get(pk = request.data['project'])
        prjT.uploadingBreakFile = False
        prjT.save()

        return Response({},status=status.HTTP_200_OK)




class DownloadPlannerView(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        workbook = Workbook()
        values =  request.GET['values']
        pkVal = values.split(',')
        pkVal = pkVal[:-1]
        Sheet1 = workbook.active
        Sheet1.title = 'Performance Report'
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        hd = ['Name','Company','TeamLead','Location','Date','TimeZone','description']
        Sheet1.append(hd)
        hdWidth = []
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            if idx == 0 or idx == len(hd)-1:
                hdWidth.append(25)
            else:
                hdWidth.append(20)
        for j in pkVal:
            plannerObj = PlannerItem.objects.get(pk=j)
            hd1=[plannerObj.name,plannerObj.company.name,plannerObj.teamLead.first_name + " " + plannerObj.teamLead.last_name,plannerObj.location,plannerObj.date, plannerObj.timeZone,plannerObj.description]
            Sheet1.append(hd1)

        for idx,i in enumerate(hdWidth):
            Sheet1.column_dimensions[alphaChars[idx]].width = i
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ProjectPlanner.xlsx'
        return response


import os
from django.http import HttpResponse
imageViewerTemplate = Template(open( os.path.join(globalSettings.BASE_DIR , 'projects', 'templates','imageViewerBase.html')).read())

def ImageViewerView(request):
    print 'image viewerrrrrrrrrrr'

    return HttpResponse(imageViewerTemplate.render(Context({'useCDN' : globalSettings.USE_CDN , 'BRAND_LOGO' : globalSettings.BRAND_LOGO ,'BRAND_NAME' :  globalSettings.BRAND_NAME})))

def ConfigurationManagerView(request):

    if not request.user.is_authenticated():
        try:
            acc = ApiAccount.objects.get(apiKey = request.GET['apikey'])
        except:
            traceback.print_exc(file=sys.stdout)
            raise SuspiciousOperation('Expired')
        login(request , acc.user, backend='django.contrib.auth.backends.ModelBackend')

    return render(request, 'configurationManagerBase.html' , {'useCDN' : globalSettings.USE_CDN , 'BRAND_LOGO' : globalSettings.BRAND_LOGO ,'BRAND_NAME' :  globalSettings.BRAND_NAME} )
    # return HttpResponse(imageViewerTemplate.render(Context({'useCDN' : globalSettings.USE_CDN , 'BRAND_LOGO' : globalSettings.BRAND_LOGO ,'BRAND_NAME' :  globalSettings.BRAND_NAME})))


def fileiterator(zipf):
	with zipfile.ZipFile(zipf, "r", zipfile.ZIP_STORED) as openzip:
		filelist = openzip.infolist()
		for f in filelist:
			yield(f.filename, openzip.read(f))

def process(zipf, callback):
	with zipfile.ZipFile(zipf, "r", zipfile.ZIP_STORED) as openzip:
		filelist = openZip.infolist()
		for f in filelist:
			callback(f.filename, openzip.read(f))

from HR.models import Team, accountsKey
import hashlib, random

@csrf_exempt
def QuickProjectView(request):
    print request.FILES , request.POST
    try:
        acc = ApiAccount.objects.get(apiKey = request.POST['apikey'])
    except:
        raise SuspiciousOperation('Expired')
    projName = request.FILES['config'].name.split('.')[0]
    folPath = os.path.join(globalSettings.BASE_DIR , 'media_root', str(acc.pk), projName)
    try:
        os.makedirs(folPath)
    except OSError as e:
        pass

    team , tcreated = Team.objects.get_or_create(title = acc.user.first_name + '_default')
    proj , created = project.objects.get_or_create(title = projName , user = acc.user, status = 'coding')

    proj.coderTeam = team
    proj.projectStart = datetime.datetime.now()
    proj.save()

    des = acc.user.designation.team.add(team)

    FileUpload.objects.filter(project = proj).delete()

    for filename,content in fileiterator(request.FILES['files']):
        f = open( os.path.join(globalSettings.BASE_DIR , 'media_root', str(acc.pk), projName ,filename), 'wb')

        FileUpload.objects.create(**{'project':proj,'path': '/media/'+ str(acc.pk) + '/'+ projName + '/'+ filename })

        f.write(content)
        f.close()

    ProjectFields.objects.filter(project = proj).delete()

    df = pd.read_excel(request.FILES['config'])
    for col in df.columns:
        typ = col[col.find("(")+1:col.find(")")]
        name = col.split('(')[0]
        p = ProjectFields.objects.create(**{'project':proj,'name': name , 'type' : typ  })
        if typ == 'name':
            p.prefix = True
            p.fName = True
            p.mName = True
            p.lName = True
            p.compName = True
            p.nameFormat = 'PN|\s|LN|,|FN|\s|MN|@|CN'
            p.save()

    salt = hashlib.sha1(str(random.random())).hexdigest()[:5]
    activation_key = hashlib.sha1(salt+acc.user.email).hexdigest()
    key_expires = datetime.datetime.today() + datetime.timedelta(2)
    accountsKey.objects.filter(user = acc.user).delete()

    ak = accountsKey(user=acc.user, activation_key=activation_key,
        key_expires=key_expires)
    ak.save()

    return JsonResponse({"url" : globalSettings.SITE_ADDRESS + '/token/?key=' + activation_key + '&next=/project/coder/' + str(proj.pk) , "pk" : proj.pk} ,status =200 )

@csrf_exempt
def downloadProjectReportByAPI(request):
    try:
        acc = ApiAccount.objects.get(apiKey = request.GET['apikey'])
        project.objects.get(user = acc.user , pk = request.GET['projectpk'])
    except:
        traceback.print_exc(file=sys.stdout)
        raise SuspiciousOperation('Expired')

    return projectReportExcel(request, True)

def getDynamicForm(request):
    df , created = DynamicForm.objects.get_or_create(user = request.user , name = request.GET['name'])
    return JsonResponse(DynamicFormSerializer(df).data ,status =200 )

def getOCRConfig(request):
    df , created = OCRConfig.objects.get_or_create(owner = request.user , name = request.GET['name'])
    return JsonResponse(OCRConfigSerializer(df).data ,status =200 )


def dynamicFormView(request):

    if not request.user.is_authenticated():
        try:
            acc = ApiAccount.objects.get(apiKey = request.GET['apikey'])
        except:
            traceback.print_exc(file=sys.stdout)
            raise SuspiciousOperation('Expired')
        login(request , acc.user, backend='django.contrib.auth.backends.ModelBackend')
    pi = DynamicForm.objects.get(name = request.GET['name'], user =request.user)
    pid = pi.pk
    if pid:
        data = {'typ':'dynamicForm','imageViewerUrl':globalSettings.IMAGEVIEWER_URL}
        data['projectid'] = pid
        data['projectname'] = "Coding : " + pi.name

        return render(request, 'projectDynamicForm.html' , data)
    else:
       return HttpResponse('', status=status.HTTP_404_NOT_FOUND)

def projectDocsView(request):

    proj = project.objects.get(pk = request.GET['project'])

    return render(request, 'projectDocs.html' , {"docs" : proj.files.all()  , "name" : proj.title  })


def allDoneView(request):
    return render(request, 'allDone.html' )

from PIL import Image as pilIMage
pilIMage.MAX_IMAGE_PIXELS = None
from pytesseract import image_to_string
from django.core.files.storage import default_storage,FileSystemStorage
from django.core.files.base import ContentFile
import fitz

@csrf_exempt
def SaveMediaFile(request):
    fileinreq = request.FILES['file']
    default_storage.save(fileinreq.name , ContentFile(fileinreq.read()))
    return JsonResponse({"filename" : fileinreq.name} ,status =200 )

@csrf_exempt
def getOCRResult(request):
    # print "files : " , request.FILES

    try:
        acc = ApiAccount.objects.get(apiKey = request.GET['apikey'])
        # acc.user
        configObj = OCRConfig.objects.filter( name = request.GET['config'] , owner = acc.user)[0]
    except:
        traceback.print_exc(file=sys.stdout)
        raise SuspiciousOperation('Expired')


    # print configObj
    grabs = json.loads(configObj.grabs)
    fileinreq = request.FILES['file']

    if fileinreq.name.endswith('tif') or fileinreq.name.endswith('tiff'):
        default_storage.save("temp__01__" + fileinreq.name , ContentFile(fileinreq.read()))

        fullImage = pilIMage.open(os.path.join(globalSettings.BASE_DIR , 'media_root' , "temp__01__" + fileinreq.name) )
    elif fileinreq.name.endswith('pdf'):
        # read the images
        default_storage.save("temp__01__" + fileinreq.name , ContentFile(fileinreq.read()))
        doc = fitz.open(os.path.join(globalSettings.BASE_DIR , 'media_root' , "temp__01__" + fileinreq.name))
        # print len(doc)
    else:
        img = pilIMage.open(fileinreq)


    toSend = {}
    for indx, grab in enumerate(grabs):
        try:
            if fileinreq.name.endswith('tif') or fileinreq.name.endswith('tiff'):
                fullImage.seek(grab['pageNumber'])
                img = fullImage.convert('RGB')
            elif fileinreq.name.endswith('pdf'):
                page = doc.loadPage(grab['pageNumber']) #number of page
                zoom_x = 3.0
                zoom_y = 3.0
                trans = fitz.Matrix(zoom_x, zoom_y).preRotate(0)
                pix = page.getPixmap(matrix=trans, alpha=False)
                mode = "RGBA" if pix.alpha else "RGB"
                img = pilIMage.frombytes(mode, [pix.width, pix.height], pix.samples)
        except:
            traceback.print_exc(file=sys.stdout)

        # print img.__class__
        # print dir(img)

        width , height = img.size
        crop = img.crop( (grab['x']*width , grab['y']*height , grab['x']*width+ grab['width']*width ,grab['y']*height + grab['height']*height) )
        txt = image_to_string(crop)
        toSend[grab['label']] = txt

    return JsonResponse(toSend ,status =200 )


@csrf_exempt
def getBulkOCRResult(request):
    configObj = OCRConfig.objects.get( name = request.GET['config'])
    # print configObj
    grabs = json.loads(configObj.grabs)
    zipFIleInReq = request.FILES['file']

    zf = zipfile.ZipFile(zipFIleInReq)
    print zf.namelist()
    toReturn = []
    for filename in zf.namelist():
        print filename
        fileinreq = zf.read(filename)

        if filename.endswith('tif') or filename.endswith('tiff'):
            f = open( os.path.join(globalSettings.BASE_DIR , 'media_root' , ('zipfile_' + filename) ) , 'wb')
            f.write(fileinreq)
            f.close()

            # fullImage = pilIMage.open(os.path.join(globalSettings.BASE_DIR , 'media_root' , "zipfile_" + filename) )
            fullImage = pilIMage.fromstring(fileinreq)
        elif filename.endswith('pdf'):
            # read the images
            f = open( os.path.join(globalSettings.BASE_DIR , 'media_root' , ('zipfile_' + filename) ) , 'wb')
            f.write(fileinreq)
            f.close()
            doc = fitz.open(os.path.join(globalSettings.BASE_DIR , 'media_root' , "zipfile_" + filename))
            print len(doc)
        else:
            img = pilIMage.fromstring(fileinreq)


        toSend = {}
        for indx, grab in enumerate(grabs):
            # try:
            if filename.endswith('tif') or filename.endswith('tiff'):
                fullImage.seek(grab['pageNumber'])
                img = fullImage.convert('RGB')
            elif filename.endswith('pdf'):
                page = doc.loadPage(grab['pageNumber']) #number of page
                zoom_x = 3.0
                zoom_y = 3.0
                trans = fitz.Matrix(zoom_x, zoom_y).preRotate(0)
                pix = page.getPixmap(matrix=trans, alpha=False)
                mode = "RGBA" if pix.alpha else "RGB"
                img = pilIMage.frombytes(mode, [pix.width, pix.height], pix.samples)
            # except:
                # traceback.print_exc(file=sys.stdout)

            # print img.__class__
            # print dir(img)

            width , height = img.size
            crop = img.crop( (grab['x']*width , grab['y']*height , grab['x']*width+ grab['width']*width ,grab['y']*height + grab['height']*height) )
            txt = image_to_string(crop)
            toSend[grab['label']] = txt
        toReturn.append(toSend)

    return JsonResponse({"data" : toReturn} ,status =200 )


class RetentionNoticeAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):

        toReturn = []
        dt = datetime.datetime.now() - datetime.timedelta(days = 90)
        for p in project.objects.filter(created__lt = dt):
            toReturn.append({"pk" : p.pk , "created" : p.created  , "name" : p.title })

        return Response(toReturn,status=status.HTTP_200_OK)
